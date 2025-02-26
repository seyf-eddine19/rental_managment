import os
import json
import openpyxl
import arabic_reshaper
from bidi.algorithm import get_display
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from decimal import Decimal

from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
from reportlab.lib.styles import ParagraphStyle

from django.conf import settings
from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone
from django.utils.text import slugify
from django.shortcuts import render, redirect, get_object_or_404

from django.contrib import messages
from django.contrib.auth import logout, update_session_auth_hash
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.forms import PasswordChangeForm

from django.http import HttpResponse
from django.urls import reverse_lazy
from django.forms import HiddenInput
from django.views.generic import ListView, FormView, DetailView, DeleteView
from django.db.models import Count, Q
from django.db import IntegrityError

from .models import Building, Apartment, Tenant, ActiveTenant, ServiceRoom
from .forms import (
    BuildingForm, ServiceRoomForm, ApartmentForm, TenantForm, ActiveTenantForm, UserForm, ProfileForm,
    BuildingFilterForm , ApartmentFilterForm, TenantFilterForm, ActiveTenantFilterForm,
    ApartmentImportForm
)

def custom_403(request, exception):
    return render(request, '403.html', {}, status=403)

def custom_404(request, exception):
    return render(request, '404.html', {}, status=404)

def custom_500(request):
    return render(request, '500.html', status=500)


@login_required
def logout_view(request):
    logout(request)
    return redirect('login')


# View for displaying and editing a user's profile
@login_required
def profile_view(request):
    user = request.user  # Get the currently logged-in user

    if request.method == 'POST':
        form = ProfileForm(request.POST, instance=user)
        password_form = PasswordChangeForm(request.user, request.POST)  # Password change form
        
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث الملف الشخصي بنجاح!')
            return redirect('profile')

        if password_form.is_valid():
            user = password_form.save()  # Save the new password
            update_session_auth_hash(request, user)  # Prevent logout
            messages.success(request, 'تم تحديث كلمة المرور بنجاح!')
            return redirect('profile')
    
    else:
        form = ProfileForm(instance=user)
        password_form = PasswordChangeForm(user)

    return render(request, 'users/profile.html', {
        'form': form,
        'password_form': password_form,
        'user': user
    })


@login_required
def index(request):
    buildings = Building.objects.all()

    building_data = []
    for building in buildings:
        apartment_counts = Apartment.objects.filter(building=building).aggregate(
            vacant=Count('id', filter=Q(status='شاغرة')),
            occupied=Count('id', filter=Q(status='مأهولة'))
        )
        building_data.append({
            'name': building.building_number,
            'vacant': apartment_counts['vacant'],
            'occupied': apartment_counts['occupied']
        })

    context = {
        'building_data_json': json.dumps(building_data, cls=DjangoJSONEncoder),
        'tenant_count': Tenant.objects.count(),
        'building_count': Building.objects.count(),
        'apartment_count': Apartment.objects.count(),
    }

    return render(request, 'index.html', context)


# Export Excels & PDFs
class ExportMixin:
    """Mixin to export any ListView data to an Excel or PDF file with proper Arabic support."""
    def get(self, request, *args, **kwargs):
        """Export data when 'export' is in request GET parameters."""
        if "export" in request.GET:
            export_format = request.GET.get("format", "excel")  # Default to 'excel'
            queryset = self.get_queryset()

            if not queryset.exists():
                return HttpResponse("No data available for export.", content_type="text/plain")

            if export_format == "pdf":
                return self.export_to_pdf(queryset)
            else:
                return self.export_to_excel(queryset)

        return super().get(request, *args, **kwargs)

    def prepare_export_response(self, content, file_type, model_name):
        """Prepare and return a file response with a timestamped filename."""
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        safe_model_name = slugify(model_name) or "export"
        filename = f"export_{safe_model_name}_{timestamp}.{file_type}"

        content_types = {
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "pdf": "application/pdf",
        }

        response = HttpResponse(content, content_type=content_types.get(file_type, "application/octet-stream"))
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def export_to_excel(self, queryset):
        """Generate an Excel file from a queryset with all fields, modern style."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"{queryset.model._meta.verbose_name_plural}"

        # Get field names dynamically
        fields = [field.name for field in queryset.model._meta.fields[1:]]
        fields1 = [field.verbose_name for field in queryset.model._meta.fields[1:]]

        # Set modern header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers with styles
        for col_num, field_name in enumerate(fields1, 1):
            cell = sheet.cell(row=1, column=col_num, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Set the column width dynamically based on the length of the content
        for col_num, field_name in enumerate(fields1, 1):
            column_width = max(len(field_name), 15)  # Ensure a minimum width of 15
            sheet.column_dimensions[get_column_letter(col_num)].width = column_width

        # Write rows
        for obj in queryset:
            row = []
            for field in fields:
                value = getattr(obj, field, "")

                # Handle related fields (ForeignKey, ManyToMany, etc.)
                if isinstance(value, str):
                    row.append(value)
                elif hasattr(value, 'get_FOO_display'):  # For choices-based fields
                    row.append(value.get_FOO_display())
                else:
                    # Check if the field is a related object (e.g., ForeignKey)
                    related_object = getattr(obj, field, None)
                    if related_object:
                        row.append(str(related_object))  # Get the string representation
                    else:
                        row.append("")

            sheet.append(row)

        # Apply border style for all cells
        border_style = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = border_style

        # Save workbook to memory
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        content = output.read()
        file_type = "xlsx"
        model_name = queryset.model._meta.model_name
        return self.prepare_export_response(content, file_type, model_name)

    def export_to_pdf(self, queryset):
        """Generate a properly formatted PDF with RTL Arabic support and custom font."""
        model_name = queryset.model._meta.model_name

        # Create an in-memory buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=10, leftMargin=10, topMargin=20, bottomMargin=20)


        # Load Arabic font
        font_path = os.path.abspath(os.path.join(settings.STATIC_ROOT, "fonts", "Janna LT Bold", "Janna LT Bold.ttf"))
        pdfmetrics.registerFont(TTFont("Janna", font_path))

        elements = []
        # Table Headers (Right-aligned for Arabic)
        fields = [[field.name, field.verbose_name] for field in queryset.model._meta.fields[-1:0:-1]]
        headers = [get_display(arabic_reshaper.reshape(field[1])) for field in fields]
        table_data = [headers]  # Table header

        max_col_lengths = [len(header) for header in headers]
        # Table Rows
        for obj in queryset:
            row = []
            for i, field in enumerate(fields):
                value = getattr(obj, field[0], "")

                if isinstance(value, bool):
                    value = "نعم" if value else "لا"
                elif isinstance(value, (float, Decimal)):
                    value = "{:,.2f}".format(value)
                elif isinstance(value, (datetime, date)):
                    value = value.strftime("%Y-%m-%d")
                elif hasattr(obj, f"get_{field[0]}_display"):
                    value = getattr(obj, f"get_{field[0]}_display")()  # Choice fields

                value = str(value) if value else ""

                # Fix Arabic text order
                if any("\u0600" <= c <= "\u06FF" for c in value):
                    value = get_display(arabic_reshaper.reshape(value))

                row.append(value)
                max_col_lengths[i] = max(max_col_lengths[i], len(value))  # Track longest text

            table_data.append(row)

        # Calculate column widths dynamically
        total_width = 780  # Approximate A4 width in landscape mode (in points)
        min_width = 80  # Minimum column width
        max_width = 220  # Maximum column width
        scale_factor = total_width / sum(max_col_lengths)  # Normalize sizes

        col_widths = [max(min_width, min(int(l * scale_factor), max_width)) for l in max_col_lengths]

        # Create Table
        table = Table(table_data, colWidths=col_widths)


        # Create Table
        # table = Table(table_data, colWidths=[len(fields)-20] * len(fields))
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Janna"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ]))

        # **Title (Centered and Bold)**
        title_text = get_display(arabic_reshaper.reshape(f"تقرير {queryset.model._meta.verbose_name_plural}"))
        title_style = ParagraphStyle(name="Title", fontName="Janna", fontSize=16, alignment=1, spaceAfter=20)

        # **Subtitle (Smaller text under the title)**
        subtitle_text = get_display(arabic_reshaper.reshape(f"قائمة {queryset.model._meta.verbose_name_plural} المصدرة من النظام"))
        subtitle_style = ParagraphStyle(name="Subtitle", fontName="Janna", fontSize=10, alignment=1, textColor=colors.grey)

        elements.append(Paragraph(title_text, title_style))
        elements.append(Spacer(1, 20))
        elements.append(table)
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(subtitle_text, subtitle_style))


        doc.build(elements)

        # Get PDF content from buffer
        pdf_content = buffer.getvalue()
        buffer.close()

        # Use `prepare_export_response` for consistent file handling
        return self.prepare_export_response(pdf_content, "pdf", model_name)


# Form Views Mixin
class FormViewMixin(FormView):
    def get_object(self):
        """Retrieve object if updating, or return None for creation."""
        pk = self.kwargs.get("pk")
        if pk:
            return get_object_or_404(self.model, pk=pk)
        return None

    def get_form_kwargs(self):
        """Pass instance to form if updating"""
        kwargs = super().get_form_kwargs()
        obj = self.get_object()
        if obj:
            kwargs["instance"] = obj
        return kwargs

    def get_initial(self):
        """Prefill form with existing data if updating."""
        obj = self.get_object()
        if obj:
            return {field.name: getattr(obj, field.name) for field in obj._meta.fields}
        return {}

    def form_valid(self, form):
        """Validate form and handle saving"""
        obj = form.save(commit=False) 
        is_update = self.get_object() is not None 

        obj.save()
        self.object = obj 
 
        if is_update:
            messages.success(self.request, "تم تحديث البيانات بنجاح!")
        else:
            messages.success(self.request, "تم إضافة البيانات بنجاح!")

        return super().form_valid(form)

    def form_invalid(self, form):
        """Handle invalid form submissions."""
        messages.error(self.request, "حدث خطأ أثناء حفظ البيانات. يرجى التحقق من المدخلات.")
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["object"] = self.get_object()
        return context


# User Views
class UserListView(PermissionRequiredMixin, ListView):
    model = User
    template_name = 'users/user_list.html'
    context_object_name = 'users'
    permission_required = 'auth.view_user'

class UserDetailView(PermissionRequiredMixin, DetailView):
    model = User
    template_name = 'users/user_detail.html'
    permission_required = 'auth.view_user'

class UserDeleteView(PermissionRequiredMixin, DeleteView):
    model = User
    template_name = 'users/user_delete.html'
    success_url = reverse_lazy('user_list')
    permission_required = 'auth.delete_user'

    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        messages.success(request, 'تم حذف البيانات بنجاح.')
        return response

class UserFormView(PermissionRequiredMixin, FormViewMixin):
    model = User
    form_class = UserForm
    template_name = 'users/user_form.html'
    success_url = reverse_lazy('user_list')
    permission_required = ['auth.add_user', 'auth.change_user']

    def dispatch(self, request, *args, **kwargs):
        """Check if we're updating an existing user"""
        self.user_instance = None
        if "pk" in kwargs:
            self.user_instance = get_object_or_404(User, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_form(self, form_class=None):
        """إرجاع النموذج مع تعيين الكائن في وضع التحديث."""
        form = super().get_form(form_class)
        obj = self.get_object()
        
        if obj:  # تحديث مستخدم موجود
            form.instance = obj
            form.fields['password'].widget = HiddenInput()
            form.fields['password'].required = False
        return form


# Building Views
class BuildingListView(PermissionRequiredMixin, ExportMixin, ListView):
    model = Building
    template_name = 'buildings/list.html'
    context_object_name = 'buildings'
    permission_required = 'rentals.view_building'

    def get_queryset(self):
        queryset = super().get_queryset()
        building_filter = self.request.GET.get('building_number')
        search_filter = self.request.GET.get('search')

        if building_filter:
            queryset = queryset.filter(building_number=building_filter)
        if search_filter:
            queryset = queryset.filter(
                Q(address__icontains=search_filter)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = BuildingFilterForm(self.request.GET)
        return context

class BuildingFormView(PermissionRequiredMixin, FormViewMixin):
    model = Building
    form_class = BuildingForm
    template_name = 'buildings/form.html'
    success_url = reverse_lazy('building_list')
    permission_required = 'rentals.add_building'  

class BuildingDetailView(PermissionRequiredMixin, DetailView):
    model = Building
    template_name = 'buildings/detail.html'
    permission_required = 'rentals.view_building'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["model_name"] = self.model._meta.model_name  
        context["model_verbose_name"] = self.model._meta.verbose_name  

        # Fetch only direct fields (exclude reverse relations)
        context["fields"] = [
            (field.verbose_name, getattr(self.object, field.name, "لا توجد بيانات"))
            for field in self.model._meta.get_fields()
            if not (field.auto_created and not field.concrete) and field.name != "id"
        ]

        context["service_rooms"] = self.object.service_rooms.all()
        
        # Check if editing a room
        room_id = self.request.GET.get("edit_room")
        if room_id:
            room = get_object_or_404(ServiceRoom, id=room_id, building=self.object)
            context['serviceroom_form'] = ServiceRoomForm(instance=room)
            context['editing_room'] = room  # Pass the room object for editing UI
        else:
            context['serviceroom_form'] = ServiceRoomForm()

        return context

    def post(self, request, *args, **kwargs):
        """Handle adding, updating, and deleting service rooms."""
        self.object = self.get_object()

        # Handle deletion
        if "delete_room" in request.POST:
            room_id = request.POST.get("delete_room")
            service_room = get_object_or_404(ServiceRoom, id=room_id, building=self.object)
            service_room.delete()
            messages.success(request, f"تم حذف غرفة الخدمات {service_room.room_number} بنجاح.")
            return redirect(self.get_success_url())

        # Handle add/update
        room_id = request.POST.get("room_id")  # Check if it's an update request
        if room_id:
            service_room = get_object_or_404(ServiceRoom, id=room_id, building=self.object)
            form = ServiceRoomForm(request.POST, instance=service_room)  # Update existing
            action = "تحديث"
        else:
            form = ServiceRoomForm(request.POST)  # Create new
            service_room = form.save(commit=False)
            service_room.building = self.object
            action = "إضافة"

        if form.is_valid():
            service_room.save()
            messages.success(request, f"تم {action} غرفة الخدمات بنجاح.")
        else:
            messages.error(request, f"حدث خطأ أثناء {action} غرفة الخدمات.")

        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy('building_detail', kwargs={'pk': self.object.pk})

class BuildingDeleteView(PermissionRequiredMixin, DeleteView):
    model = Building
    template_name = 'buildings/delete.html'
    success_url = reverse_lazy('building_list')
    permission_required = 'rentals.delete_building'

    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        messages.success(request, 'تم حذف البيانات بنجاح.')
        return response


# Apartment Views
class ApartmentListView(PermissionRequiredMixin, ExportMixin, ListView):
    model = Apartment
    template_name = 'apartments/list.html'
    context_object_name = 'apartments'
    permission_required = 'rentals.view_apartment'

    def get_queryset(self):
        queryset = super().get_queryset()
        building_filter = self.request.GET.get('building')
        apartment_filter = self.request.GET.get('apartment')
        status_filter = self.request.GET.get('status')
        search_filter = self.request.GET.get('search')

        if building_filter:
            queryset = queryset.filter(building__building_number=building_filter)
        if apartment_filter:
            queryset = queryset.filter(apartment_number=apartment_filter)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if search_filter:
            queryset = queryset.filter(
                Q(electricity_meter_number__icontains=search_filter) |
                Q(water_meter_number__icontains=search_filter)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ApartmentFilterForm(self.request.GET)
        context['import_form'] = ApartmentImportForm()
        return context

    def post(self, request, *args, **kwargs):
        """استيراد الشقق من ملف Excel"""
        form = ApartmentImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["import_file"]

            if not excel_file.name.endswith(".xlsx"):
                messages.error(request, "يرجى تحميل ملف Excel بصيغة (.xlsx) فقط.")
                return redirect("apartment_list")

            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active  # الحصول على الورقة النشطة

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        building_number, apartment_number, num_of_rooms, electricity_meter, water_meter, floor_number = row

                        if not all([building_number, apartment_number, num_of_rooms, electricity_meter, water_meter, floor_number]):
                            messages.error(request, f"بيانات ناقصة في أحد الصفوف، يرجى التحقق من الملف.")
                            continue

                        building, created = Building.objects.get_or_create(
                            building_number=building_number,
                            defaults={"number_of_floors": 5, "number_of_apartments": 0}
                        )

                        Apartment.objects.create(
                            building=building,
                            apartment_number=apartment_number,
                            num_of_rooms=int(num_of_rooms),
                            electricity_meter_number=electricity_meter,
                            water_meter_number=water_meter,
                            status='شاغرة',
                            floor_number=int(floor_number),
                        )

                        building.number_of_apartments = building.apartments.count()
                        building.save()

                    except Exception as e:
                        messages.error(request, f"حدث خطأ أثناء استيراد البيانات: {building}|{apartment_number} \n {e}")

                messages.success(request, "تم استيراد الشقق بنجاح.")
                return redirect("apartment_list")

            except Exception as e:
                messages.error(request, f"تعذر قراءة ملف Excel: {e}")
                return redirect("apartment_list")

        return redirect("apartment_list")

class ApartmentFormView(PermissionRequiredMixin, FormViewMixin):
    model = Apartment
    form_class = ApartmentForm
    template_name = "apartments/form.html"
    success_url = reverse_lazy('apartment_list')
    permission_required = 'rentals.add_apartment'
    
    def get_success_url(self):
        return reverse_lazy('apartment_detail', kwargs={'pk': self.object.pk})

class ApartmentDetailView(PermissionRequiredMixin, DetailView):
    model = Apartment
    template_name = 'apartments/detail.html'
    permission_required = 'rentals.view_apartment'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["model_name"] = self.model._meta.model_name  # Model name for template
        context["model_verbose_name"] = self.model._meta.verbose_name  # Model name for template

        # Fetch only direct fields (exclude reverse relations)
        context["fields"] = [
            (field.verbose_name, getattr(self.object, field.name, "لا توجد بيانات"))
            for field in self.model._meta.get_fields()
            if not (field.auto_created and not field.concrete) and field.name != "id"
        ]
        return context
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        context = self.get_context_data()

        # Handle export request
        if "export" in self.request.GET:
            export_format = self.request.GET.get("format", "excel")
            if export_format == "pdf":
                return self.export_to_pdf()
        return self.render_to_response(context)
    
    def post(self, request, *args, **kwargs):
        """ Handle the 'Set Vacant' action directly in this view. """
        apartment = self.get_object()

        if "set_vacant" in request.POST:  
            active_tenant = apartment.active_tenant 
            if active_tenant:
                active_tenant.delete()
                messages.success(request, "تم تعيين الشقة كشاغرة بنجاح.")
            else:
                messages.warning(request, "الشقة بالفعل شاغرة.")

        return redirect('apartment_detail', pk=apartment.pk)

    def export_to_pdf(self):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    
        # Register Arabic font
        font_path = os.path.abspath(os.path.join(settings.STATIC_ROOT, "fonts", "Janna LT Bold", "Janna LT Bold.ttf"))
        pdfmetrics.registerFont(TTFont("Janna", font_path))
    
        # Styles
        title_style = ParagraphStyle(name="Title", fontName="Janna", fontSize=16, alignment=2, spaceAfter=20, textColor=colors.HexColor("#336699"))
        subtitle_style = ParagraphStyle(name="Subtitle", fontName="Janna", fontSize=12, alignment=2, spaceAfter=20, textColor=colors.black)
        text_style = ParagraphStyle(name="Text", fontName="Janna", fontSize=10, alignment=2, spaceAfter=5)  # Right-aligned for Arabic text
        table_style = TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, -1), "Janna"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("LINEABOVE", (0, 0), (-1, 1), 0.5, colors.grey),  # Border above the first row
            ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.grey),  # Border below the last row
       ])

        # Prepare content for PDF export, adjust as needed for your model
        text_1 = (
          get_display(arabic_reshaper.reshape("رقم العمارة: ") + "<font color='darkgray'>" + str(self.object.building.building_number) + "</font>")
        )
        text_2 = (
          get_display(arabic_reshaper.reshape("رقم الشقة: ") + "<font color='darkgray'>" + str(self.object.apartment_number) + "</font>")
        )
        apartment_details = [
            [
             Paragraph(text_2, subtitle_style),
             Paragraph(text_1, subtitle_style)
            ], [
             Paragraph(get_display(arabic_reshaper.reshape("عدد الغرف: ") + "<font color='darkgray'>" + str(self.object.num_of_rooms) + "</font>"), subtitle_style), 
             Paragraph("<font color='darkgray'>" + get_display(arabic_reshaper.reshape(str(self.object.building.address))) + "</font>" + get_display(arabic_reshaper.reshape("عنوان العمارة: ")), subtitle_style), 
            ], [
             Paragraph(get_display(arabic_reshaper.reshape("رقم عداد الماء: ") + "<font color='darkgray'>" + str(self.object.water_meter_number) + "</font>"), subtitle_style),
             Paragraph(get_display(arabic_reshaper.reshape("رقم عداد الكهرباء: ") + "<font color='darkgray'>" + str(self.object.electricity_meter_number) + "</font>"), subtitle_style),
            ], [ 
             Paragraph("<font color='darkgray'>" + get_display(arabic_reshaper.reshape(str(self.object.status))) + "</font>" + get_display(arabic_reshaper.reshape("حالة الشقة: ")), subtitle_style), 
             Paragraph(get_display(arabic_reshaper.reshape("رقم الطابق: ") + "<font color='darkgray'>" + str(self.object.floor_number) + "</font>"), subtitle_style), 
            ]
        ]

        # Tenant information (you can adjust the model references accordingly)
        tenant_details = [
            [
             Paragraph(f"<font color='darkgray'>{self.object.active_tenant.tenant.phone_number}</font> <strong>{get_display(arabic_reshaper.reshape('رقم الهاتف: '))}</strong> ", subtitle_style),
             Paragraph(f"<font color='darkgray'>{get_display(arabic_reshaper.reshape(self.object.active_tenant.tenant.name))}</font> <strong>{get_display(arabic_reshaper.reshape('اسم المستأجر: '))}</strong> ", subtitle_style),
            ], [
             Paragraph(f"<font color='darkgray'>{self.object.active_tenant.rent_amount}</font> <strong>{get_display(arabic_reshaper.reshape('مبلغ الإيجار: '))}</strong> ", subtitle_style),
             Paragraph(f"<font color='darkgray'>{self.object.active_tenant.contract_number}</font> <strong>{get_display(arabic_reshaper.reshape('رقم العقد: '))}</strong>", subtitle_style),
            ], [
             Paragraph(f"<font color='darkgray'>{self.object.active_tenant.contract_end_date}</font> <strong>{get_display(arabic_reshaper.reshape('تاريخ انتهاء العقد:'))}</strong> ", subtitle_style),
             Paragraph(f"<font color='darkgray'>{self.object.active_tenant.contract_start_date}</font> <strong>{get_display(arabic_reshaper.reshape('تاريخ بداية العقد: '))}</strong> ", subtitle_style),
            ]
        ]
        
        rental_history = self.object.rental_history.all()
        rental_history = self.object.rental_history.values('contract_number', 'tenant__name', 'contract_start_date', 'contract_end_date', 'rent_amount', 'notes')
        
        # Prepare content for rental history
        rental_history_content = []
        for rental in rental_history:
            rental_text = f"""
            <font color='darkgrey'>{rental['rent_amount']}</font> {get_display(arabic_reshaper.reshape("بمبلغ: "))} <font color='darkgrey'>{rental['contract_end_date']}</font> {get_display(arabic_reshaper.reshape("إلى"))} <font color='darkgrey'>{rental['contract_start_date']}</font> {get_display(arabic_reshaper.reshape("من"))}  <font color='darkgrey'>{get_display(arabic_reshaper.reshape(rental['tenant__name']))}</font> {get_display(arabic_reshaper.reshape("تم تاجير الشقة ل "))} - <font color='darkgrey'> {rental['contract_number']} {get_display(arabic_reshaper.reshape("العقد: "))}</font>   
            """
            rental_history_content.append([Paragraph((rental_text), text_style)])

        elements = [
            Spacer(1, 10),
            Paragraph(get_display(arabic_reshaper.reshape(f"تفاصيل الشقة:")), title_style),
            Table(apartment_details),
            Spacer(1, 30),
            Paragraph(get_display(arabic_reshaper.reshape(f"معلومات المستاجر الحالي")), title_style),
            Table(tenant_details),
            Spacer(1, 30),
            Paragraph(get_display(arabic_reshaper.reshape("سجل الإيجار")), ParagraphStyle(name="Subtitle", fontName="Janna", fontSize=12, alignment=2, spaceAfter=20, textColor=colors.HexColor("#336699"))),
            Table(rental_history_content, style=table_style),
            Spacer(1, 10),
        ]

        # Build the PDF document
        doc.build(elements)

        # Send the PDF response
        pdf = buffer.getvalue()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.object}_details.pdf"'
        return response

class ApartmentDeleteView(PermissionRequiredMixin, DeleteView):
    model = Apartment
    template_name = 'apartments/delete.html'
    success_url = reverse_lazy('apartment_list')
    permission_required = 'rentals.delete_apartment'

    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        messages.success(request, 'تم حذف البيانات بنجاح.')
        return response


# Tenant Views
class TenantListView(PermissionRequiredMixin, ExportMixin, ListView):
    model = Tenant
    template_name = 'tenants/list.html'
    context_object_name = 'tenants'
    permission_required = 'rentals.view_tenant'

    def get_queryset(self):
        queryset = super().get_queryset()
        search_filter = self.request.GET.get('search')
        if search_filter:
            queryset = queryset.filter(
                Q(name__icontains=search_filter) |
                Q(phone_number__icontains=search_filter) |
                Q(id_number__icontains=search_filter) |
                Q(workplace__icontains=search_filter)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = TenantFilterForm(self.request.GET)
        return context

class TenantFormView(PermissionRequiredMixin, FormViewMixin):
    model = Tenant
    form_class = TenantForm
    template_name = 'tenants/form.html'
    success_url = reverse_lazy('tenant_list')
    permission_required = 'rentals.add_tenant'

class TenantDetailView(PermissionRequiredMixin, DetailView):
    model = Tenant
    template_name = 'tenants/detail.html'
    permission_required = 'rentals.view_tenant'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["model_name"] = self.model._meta.model_name  # Model name for template
        context["model_verbose_name"] = self.model._meta.verbose_name  # Model name for template

        # Fetch only direct fields (exclude reverse relations)
        context["fields"] = [
            (field.verbose_name, getattr(self.object, field.name, "لا توجد بيانات"))
            for field in self.model._meta.get_fields()
            if not (field.auto_created and not field.concrete) and field.name != "id"
        ]
        return context
    
class TenantDeleteView(PermissionRequiredMixin, DeleteView):
    model = Tenant
    template_name = 'tenants/delete.html'
    success_url = reverse_lazy('tenant_list')
    permission_required = 'rentals.delete_tenant'

    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        messages.success(request, 'تم حذف البيانات بنجاح.')
        return response


# Active Tenant Views 
class ActiveTenantListView(PermissionRequiredMixin, ExportMixin, ListView):
    model = ActiveTenant
    template_name = 'active_tenants/list.html'
    context_object_name = 'active_tenants'
    permission_required = 'rentals.view_activetenant'

    def get_queryset(self):
        queryset = super().get_queryset()
        apartment_filter = self.request.GET.get('apartment')
        building_filter = self.request.GET.get('building')
        tenant_filter = self.request.GET.get('tenant')
        search_filter = self.request.GET.get('search')

        if apartment_filter:
            queryset = queryset.filter(apartment__apartment_number=apartment_filter)
        if building_filter:
            queryset = queryset.filter(apartment__building__building_number=building_filter)
        if tenant_filter:
            queryset = queryset.filter(tenant__name=tenant_filter)
        if search_filter:
            queryset = queryset.filter(
                Q(tenant__phone_number__icontains=search_filter) |
                Q(contract_number__icontains=search_filter)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ActiveTenantFilterForm(self.request.GET)
        return context

class ActiveTenantFormView1(PermissionRequiredMixin, FormView):
    """ View to create or update an ActiveTenant """
    model = ActiveTenant
    form_class = ActiveTenantForm
    template_name = 'active_tenants/form.html'
    permission_required = 'rentals.change_apartment'

    def get_form_kwargs(self):
        """ Pass instance for update; create new if none exists """
        kwargs = super().get_form_kwargs()
        self.apartment = get_object_or_404(Apartment, id=self.kwargs['apartment_id'])

        if 'pk' in self.kwargs:
            self.active_tenant = get_object_or_404(ActiveTenant, id=self.kwargs['pk'])
        else:
            self.active_tenant = None  # Create new tenant

        kwargs['instance'] = self.active_tenant
        return kwargs

    def get_context_data(self, **kwargs):
        """ Pass apartment and active tenant IDs to the template """
        context = super().get_context_data(**kwargs)
        context['apartment_id'] = self.apartment.pk
        context['active_tenant_id'] = self.active_tenant.pk if self.active_tenant else None
        context['tenant_form'] = TenantForm()  # Render the tenant form as well
        return context
    
    def form_valid(self, form):
        """ Save the tenant details """
        # Check if a tenant was selected or not
        tenant_form = TenantForm(self.request.POST)
        
        if tenant_form.is_valid():
            tenant_name = tenant_form.cleaned_data['tenant_name']
            phone_number = tenant_form.cleaned_data['phone_number']
            id_number = tenant_form.cleaned_data['id_number']
            workplace = tenant_form.cleaned_data['workplace']

            # Handle tenant creation if not selected
            tenant, created = Tenant.objects.get_or_create(
                phone_number=phone_number,
                id_number=id_number,
                defaults={'name': tenant_name, 'workplace': workplace}
            )
            if created:
                messages.success(self.request, f"تم إضافة المستأجر الجديد: {tenant_name}")
            else:
                messages.info(self.request, f"المستأجر {tenant_name} موجود بالفعل.")
            
            # Save the active tenant instance
            form.instance.tenant = tenant
        else:
            messages.error(self.request, "هناك خطأ في بيانات المستأجر.")

        # Set the apartment for this active tenant
        form.instance.apartment = self.apartment
        form.instance.save()

        messages.success(self.request, "تم حفظ البيانات بنجاح.")
        return super().form_valid(form)

    def get_success_url(self):
        """ Redirect to apartment detail page """
        return reverse_lazy('apartment_detail', kwargs={'pk': self.apartment.pk})

class ActiveTenantFormView(PermissionRequiredMixin, FormView):
    """ View to create or update an ActiveTenant """
    model = ActiveTenant
    form_class = ActiveTenantForm
    template_name = 'active_tenants/form.html'
    permission_required = 'rentals.change_apartment'
    
    def get_form_kwargs(self):
        """ Pass instance for update; create new if none exists """
        kwargs = super().get_form_kwargs()
        self.apartment = get_object_or_404(Apartment, id=self.kwargs['apartment_id'])

        # Retrieve the ActiveTenant instance if it exists
        if 'pk' in self.kwargs:
            self.active_tenant = get_object_or_404(ActiveTenant, id=self.kwargs['pk'])
            # Populate the tenant data in the form for update
            kwargs['initial'] = {
                'tenant_name': self.active_tenant.tenant.name,
                'phone_number': self.active_tenant.tenant.phone_number,
                'id_number': self.active_tenant.tenant.id_number,
                'workplace': self.active_tenant.tenant.workplace,
                'tenant': self.active_tenant.tenant,  # Set the current tenant in the form
            }
        else:
            self.active_tenant = None  # Creating a new tenant

        kwargs['instance'] = self.active_tenant
        return kwargs

    def get_context_data(self, **kwargs):
        """ Pass apartment and active tenant IDs to the template """
        context = super().get_context_data(**kwargs)
        context['apartment_id'] = self.apartment.pk
        context['active_tenant_id'] = self.active_tenant.pk if self.active_tenant else None
        return context
    
    def form_valid(self, form):
        tenant_name = form.cleaned_data['tenant_name']
        phone_number = form.cleaned_data['phone_number']
        id_number = form.cleaned_data['id_number']
        workplace = form.cleaned_data['workplace']

        try:
            # Handle tenant creation or selection
            tenant, created = Tenant.objects.get_or_create(
                phone_number=phone_number,
                id_number=id_number,
                defaults={
                        'name': tenant_name,
                        'workplace': workplace,
                }
            )
            
            if created:
                messages.success(self.request, f"تم إضافة المستأجر الجديد: {tenant_name}")
            else:
                messages.info(self.request, f"المستأجر {tenant_name} موجود بالفعل.")
    
        except IntegrityError:
            messages.error(self.request, "رقم الهاتف أو رقم الهوية غير فريد. الرجاء التحقق من البيانات.")
            return self.form_invalid(form)
        
        # Save the active tenant instance
        form.instance.tenant = tenant
        form.instance.apartment = self.apartment
        form.instance.save()

        # Optionally, update apartment status (e.g., set as occupied)
        self.apartment.status = 'مأهولة'
        self.apartment.save()

        messages.success(self.request, "تم حفظ البيانات بنجاح.")
        return super().form_valid(form)

    def get_success_url(self):
        """ Redirect to apartment detail page """
        return reverse_lazy('apartment_detail', kwargs={'pk': self.apartment.pk})

